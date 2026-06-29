"""Named-table I/O over the .xlsx workbooks (openpyxl).

Tables are accessed by their declared name (tblPeople, tblGeo, ...).
Appends extend the table's ref so Excel keeps treating it as one table.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


class Workbook:
    def __init__(self, path: Path):
        self.path = Path(path)
        self.wb = load_workbook(self.path)
        self._index = self._build_index()

    def _build_index(self) -> dict[str, Any]:
        idx = {}
        for ws in self.wb.worksheets:
            for name, tab in ws.tables.items():
                idx[name] = ws
        return idx

    # --- read ---------------------------------------------------------
    def headers(self, table: str) -> list[str]:
        ws = self._index[table]
        c0, r0, c1, _ = range_boundaries(ws.tables[table].ref)
        return [ws.cell(row=r0, column=c).value for c in range(c0, c1 + 1)]

    def rows(self, table: str) -> list[dict]:
        ws = self._index[table]
        c0, r0, c1, r1 = range_boundaries(ws.tables[table].ref)
        hdr = self.headers(table)
        out = []
        for r in range(r0 + 1, r1 + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(c0, c1 + 1)]
            if all(v is None for v in vals):
                continue
            out.append(dict(zip(hdr, vals)))
        return out

    # --- append -------------------------------------------------------
    def append_rows(self, table: str, records: Iterable[list]) -> int:
        ws = self._index[table]
        tab = ws.tables[table]
        c0, r0, c1, r1 = range_boundaries(tab.ref)
        # If table is header-only, r1 == r0.
        next_row = r1 + 1 if r1 > r0 or self._has_data(ws, c0, r0, c1) else r0 + 1
        n = 0
        for rec in records:
            for j, val in enumerate(rec):
                ws.cell(row=next_row, column=c0 + j, value=val)
            next_row += 1
            n += 1
        if n:
            last = next_row - 1
            tab.ref = f"{get_column_letter(c0)}{r0}:{get_column_letter(c1)}{last}"
        return n

    @staticmethod
    def _has_data(ws, c0, r0, c1) -> bool:
        return any(ws.cell(row=r0 + 1, column=c).value is not None
                   for c in range(c0, c1 + 1))

    # --- meta_kv key/value table -------------------------------------
    def get_meta(self, key: str, default=None):
        for row in self.rows("meta_kv"):
            if row.get("key") == key:
                return row.get("value", default)
        return default

    def set_meta(self, key: str, value) -> None:
        ws = self._index["meta_kv"]
        tab = ws.tables["meta_kv"]
        c0, r0, c1, r1 = range_boundaries(tab.ref)
        hdr = self.headers("meta_kv")
        kcol = c0 + hdr.index("key")
        vcol = c0 + hdr.index("value")
        for r in range(r0 + 1, r1 + 1):
            if ws.cell(row=r, column=kcol).value == key:
                ws.cell(row=r, column=vcol, value=value)
                return
        self.append_rows("meta_kv", [[key if h == "key" else
                                      (value if h == "value" else None)
                                      for h in hdr]])

    # --- upsert by key column (for tblFactions writeback) ------------
    def upsert(self, table: str, key_col: str, key_val, updates: dict) -> None:
        ws = self._index[table]
        tab = ws.tables[table]
        c0, r0, c1, r1 = range_boundaries(tab.ref)
        hdr = self.headers(table)
        kidx = c0 + hdr.index(key_col)
        for r in range(r0 + 1, r1 + 1):
            if ws.cell(row=r, column=kidx).value == key_val:
                for col, v in updates.items():
                    ws.cell(row=r, column=c0 + hdr.index(col), value=v)
                return
        # not found -> append
        rec = [updates.get(h, (key_val if h == key_col else None)) for h in hdr]
        self.append_rows(table, [rec])

    def save(self) -> None:
        self.wb.save(self.path)
    
    