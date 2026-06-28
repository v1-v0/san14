"""
build_skeleton.py
Generates the structural skeleton for the SAN14 recruitment pipeline.

Requires: openpyxl  (pip install openpyxl)
Run:       python build_skeleton.py
Output:    san14_skeleton.xlsx   (a FRESH structure; see safety guard)

NOTE: This writes a *skeleton* with small seed rows. Your live, populated
      workbook is san14_dict.xlsx (1000 people / 218 places) — this script
      will NOT touch it. It writes to san14_skeleton.xlsx and aborts
      if that file already exists, so real data can never be clobbered.
"""

import os
import sys
import json
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "san14_skeleton.xlsx"
SCHEMA_VERSION = 4   # bumped: +appoint/besiege/explore patterns;
                     #         events gained parent_location, issuer, office;
                     #         new color_roles contract sheet

# ----------------------------------------------------------------------
# Safety guard — never overwrite an existing (possibly populated) file
# ----------------------------------------------------------------------
if os.path.exists(OUTPUT):
    sys.exit(
        f"✗ refusing to overwrite existing '{OUTPUT}'.\n"
        f"  Delete/rename it first if you really want a fresh skeleton.\n"
        f"  (Your live data file san14_dict.xlsx is never touched by this script.)"
    )

# ----------------------------------------------------------------------
# Styling helpers
# ----------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "A2"


def make_table(ws, name, headers, rows, widths=None,
               style="TableStyleMedium2", empty_row=False):
    """
    Write headers + rows and register an Excel ListObject (named table).

    An Excel table's ref must span at least one data row. For sheets that
    should start empty (events / pending / log_raw), pass empty_row=True to
    emit a single BLANK data row — nothing to delete later, just start typing.
    """
    ws.append(headers)
    if not rows and empty_row:
        rows = [[None] * len(headers)]
    for r in rows:
        ws.append(r)

    last_row = max(len(rows) + 1, 2)
    last_col = get_column_letter(len(headers))
    ref = f"A1:{last_col}{last_row}"
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)
    style_header(ws, len(headers))
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return last_row


def slots(**kw):
    return json.dumps(kw, ensure_ascii=False)


wb = Workbook()

# ======================================================================
# 1) people  ->  tblPeople
#    ⚠️ names are NOT unique (e.g. 李豐 ×3). `id` is the real key.
# ======================================================================
ws = wb.active
assert ws is not None          # a fresh Workbook() always has an active sheet
ws.title = "people"
people_seed = [
    [1, "曹操",   "=LEN(B2)", True],
    [2, "劉備",   "=LEN(B3)", True],
    [3, "孫權",   "=LEN(B4)", True],
    [4, "諸葛亮", "=LEN(B5)", True],
    [5, "司馬懿", "=LEN(B6)", True],
    [6, "于禁",   "=LEN(B7)", True],
    [7, "王淩",   "=LEN(B8)", True],
    [8, "獻帝",   "=LEN(B9)", True],   # NEW: appoint issuer (court figure)
]
make_table(
    ws, "tblPeople",
    ["id", "name", "name_len", "active"],
    people_seed,
    widths=[8, 16, 12, 10],
)

# ======================================================================
# 2) geo  ->  tblGeo
#    ⚠️ place names span len 1–3 (single-char 據點: 薊·吳·鄴·郯·魯…)
# ======================================================================
ws = wb.create_sheet("geo")
geo_seed = [
    ["薊",   "都市", "幽州", "燕國",   "=LEN(A2)", "len-1 example"],
    ["許昌", "都市", "豫州", "潁川郡", "=LEN(A3)", ""],
    ["成都", "都市", "益州", "蜀郡",   "=LEN(A4)", ""],
    ["虎牢關", "關隘", "司州", "河南尹", "=LEN(A5)", "len-3 example"],
    ["平原", "都市", "青州", "平原國", "=LEN(A6)", "parent_location example"],
    ["高唐", "都市", "青州", "平原國", "=LEN(A7)", "child of 平原"],
    ["街亭", "關隘", "雍州", "天水郡", "=LEN(A8)", "besiege example"],
]
make_table(
    ws, "tblGeo",
    ["據點名稱", "據點類別", "所屬州", "所屬郡國", "name_len", "備注"],
    geo_seed,
    widths=[14, 12, 10, 12, 12, 30],
)

# ======================================================================
# 3) patterns  ->  tblPatterns  (Stage-4 grammar registry, priority-ordered)
#    NEW: appoint(15), besiege(32), explore(50)
#    MOD: capture now accepts [軍隊] + optional (所屬於<parent>)
# ======================================================================
ws = wb.create_sheet("patterns")
patterns_seed = [
    ["recruit",  r"^(?P<subject>.{2,4}?)登庸(?P<object>.{2,4}?)(?P<result>成功|失敗)$",
     slots(subject="people", object="people"), 10, True, "劉備登庸王淩成功"],
    ["appoint",  r"^(?P<subject>.{2,4}?)接受(?P<issuer>.{1,4}?)敕令[，,]就任(?P<office>.{2,6}?)$",
     slots(subject="people", issuer="people", office="raw"), 15, True, "士燮接受獻帝敕令，就任州刺史"],
    ["promote",  r"^(?P<object>.{2,4}?)就任(?P<subject>.{2,4}?)勢力君主$",
     slots(subject="people", object="people"), 20, True, "曹操就任劉備勢力君主"],
    ["destroy",  r"^(?P<target_faction>.{2,4}?)軍(?:勢力)?滅(?:止|亡)$",
     slots(target_faction="people"), 20, True, "袁紹軍滅止"],
    ["capture",  r"^(?P<subject>.{2,4}?)[軍隊](?:攻陷|攻略|占領)(?P<location>.{1,3}?)(?:[（(]所屬於(?P<parent>.{1,3}?)[）)])?$",
     slots(subject="people", location="places", parent="places"), 30, True, "張飛隊占領高唐(所屬於平原)"],
    ["besiege",  r"^(?:(?P<subject>.{2,4}?)[軍隊])?已?包圍(?P<location>.{1,3}?)$",
     slots(subject="people", location="places"), 32, True, "已包圍街亭"],
    ["march",    r"^(?P<subject>.{2,4}?)隊(?:向|往)?(?P<location>.{1,3}?)出陣$",
     slots(subject="people", location="places"), 40, True, "于禁隊向許昌出陣"],
    ["explore",  r"^(?P<subject>.{2,4}?)探索(?P<location>.{1,3}?)(?P<result>成功|失敗)$",
     slots(subject="people", location="places"), 50, True, "崔琰探索平原失敗"],
]
make_table(
    ws, "tblPatterns",
    ["action_type", "regex", "slots", "priority", "confirmed", "example"],
    patterns_seed,
    widths=[14, 58, 40, 10, 12, 26],
)
for col in ("B", "C"):
    for cell in ws[col]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# ======================================================================
# 4) color_roles  ->  tblColorRoles  (gate contract; HSV numbers live in dict)
#    Cross-checked AFTER regex parse. The skeleton holds the RULE
#    (location must be blue, else demote); san14_dict holds the
#    MEASUREMENT (blue = these H/S/V ranges). Recalibration touches dict only.
# ======================================================================
ws = wb.create_sheet("color_roles")
color_roles_seed = [
    ["location",        "blue",       "demote",  "→ tblPending"],
    ["parent_location", "blue",       "demote",  "capture only"],
    ["subject_own",     "green",      "demote",  "own / allied actor"],
    ["subject_hostile", "red",        "demote",  "hostile actor"],
    ["target_faction",  "red",        "demote",  "destroy"],
    ["issuer",          "red|white",  "warn",    "court ≠ player faction; keep on mismatch"],
    ["office",          "any",        "ignore",  "raw slot; NEVER a dedup key or anchor"],
]
make_table(
    ws, "tblColorRoles",
    ["slot", "expected_color", "on_mismatch", "note"],
    color_roles_seed,
    widths=[18, 16, 14, 36],
)

# ======================================================================
# 5) factions  ->  tblFactions  (ruler state; subset of people)
# ======================================================================
ws = wb.create_sheet("factions")
factions_seed = [
    ["曹操", True,  "190年1月",     ""],
    ["劉備", True,  "190年1月",     ""],
    ["孫權", True,  "190年1月",     ""],
    ["袁紹", False, "193年2月下旬", "滅止"],
]
make_table(
    ws, "tblFactions",
    ["name", "is_ruler", "since_date", "note"],
    factions_seed,
    widths=[16, 12, 16, 24],
)

# ======================================================================
# 6) events  ->  tblEvents  (Stage-5 master output; starts EMPTY)
#    NEW columns vs schema v3:
#      issuer          — appoint (people-verified)
#      parent_location — capture; PERSISTED per decision (places-verified)
#      office          — appoint; RAW, NOT vocab-verified, NEVER a dedup key
#    Retained: ambiguous_id (set when a `people` slot resolves to >1 id)
# ======================================================================
ws = wb.create_sheet("events")
events_headers = [
    "date", "turn_period", "ruler", "action_type", "subject", "object",
    "issuer", "ambiguous_id", "location", "parent_location",
    "target_faction", "office", "value", "success", "confidence", "raw_text",
]
make_table(
    ws, "tblEvents", events_headers, [], empty_row=True,
    widths=[12, 10, 10, 12, 10, 10, 10, 12, 10, 14, 12, 12, 8, 9, 11, 30],
)

# ======================================================================
# 7) pending  ->  tblPending  (Stage-4 unmatched / low-conf queue; EMPTY)
# ======================================================================
ws = wb.create_sheet("pending")
make_table(
    ws, "tblPending",
    ["screenshot_id", "raw_text", "skeleton", "cluster_size",
     "best_score", "suggested_action", "resolved"],
    [], empty_row=True,
    widths=[14, 30, 22, 12, 10, 16, 10],
)

# ======================================================================
# 8) log_raw  ->  tblLogRaw  (raw OCR audit trail; EMPTY)
# ======================================================================
ws = wb.create_sheet("log_raw")
make_table(
    ws, "tblLogRaw",
    ["screenshot_id", "roi", "line_no", "raw_line", "captured_at"],
    [], empty_row=True,
    widths=[14, 16, 9, 32, 22],
)

# ======================================================================
# 9) meta  ->  meta_kv  (run state / schema + derived counts)
# ======================================================================
ws = wb.create_sheet("meta")
make_table(
    ws, "meta_kv",
    ["key", "value"],
    [
        ["schema_version",    SCHEMA_VERSION],
        ["last_screenshot",   "0000"],
        ["people_count",      "=COUNTA(tblPeople[name])"],
        ["places_count",      "=COUNTA(tblGeo[據點名稱])"],
        ["faction_count",     "=COUNTIF(tblFactions[is_ruler], TRUE)"],
        ["pattern_count",     "=COUNTA(tblPatterns[action_type])"],
        ["people_len_buckets", "2-4"],
        ["places_len_buckets", "1-3"],
    ],
    widths=[22, 26],
)

# ----------------------------------------------------------------------
# Data validation (dropdowns) — applied AFTER sheets exist
# ----------------------------------------------------------------------
def add_bool_dv(ws, col_letter, first=2, last=5000):
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first}:{col_letter}{last}")

add_bool_dv(wb["people"],   "D")   # active
add_bool_dv(wb["factions"], "B")   # is_ruler
add_bool_dv(wb["patterns"], "E")   # confirmed
add_bool_dv(wb["events"],   "N")   # success  (shifted: +issuer/parent_location/office)
add_bool_dv(wb["pending"],  "G")   # resolved

# action_type dropdown on events, sourced from the patterns column
dv_action = DataValidation(
    type="list", formula1="=patterns!$A$2:$A$500", allow_blank=True
)
wb["events"].add_data_validation(dv_action)
dv_action.add("D2:D5000")

# on_mismatch dropdown on color_roles (governs gate behaviour)
dv_mismatch = DataValidation(
    type="list", formula1='"demote,warn,ignore"', allow_blank=True
)
wb["color_roles"].add_data_validation(dv_mismatch)
dv_mismatch.add("C2:C500")

# ----------------------------------------------------------------------
# Sheet order & save (public move_sheet API, front-to-back insertion)
# ----------------------------------------------------------------------
order = ["meta", "people", "geo", "patterns", "color_roles", "factions",
         "events", "pending", "log_raw"]
for target_idx, name in enumerate(order):
    cur_idx = wb.sheetnames.index(name)
    if cur_idx != target_idx:
        wb.move_sheet(name, offset=target_idx - cur_idx)

wb.save(OUTPUT)
print(f"✓ {OUTPUT} created — 9 sheets, 9 named tables, schema v{SCHEMA_VERSION}.")
print("  events / pending / log_raw start with one blank table row (just type in it).")