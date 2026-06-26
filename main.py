"""
Parse a 三國志14 (Sangokushi 14) action log and produce two shortlists,
then export the original data, detected OCR/merge issues, and both
shortlists to a new multi-sheet .xlsx file.

  Sheet 1 (Original)            - the raw Actions column as ingested.
  Sheet 2 (Issues)              - rows that needed OCR normalization or splitting.
  Sheet 3 (Shortlist1_Contested)- contested 中止登庸 targets, counted.
  Sheet 4 (Shortlist2_Outcomes) - full acquisition outcome per target.
"""
import csv
import os
import re
from collections import defaultdict, Counter


# ---------------------------------------------------------------------------
# 1. Load the "Actions" column
# ---------------------------------------------------------------------------
def load_actions(path: str) -> list[str]:
    """Read the 'Actions' column from an .xlsx (preferred) or .csv file."""
    if path.lower().endswith(".xlsx"):
        import openpyxl  # pip install openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("No active worksheet found in the workbook.")

        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return []

        header = [str(c).strip() if c is not None else "" for c in header_row]
        try:
            col = header.index("Actions")
        except ValueError:
            col = len(header) - 1  # fall back to last column

        rows = []
        for r in rows_iter:
            if r and col < len(r) and r[col]:
                rows.append(str(r[col]).strip())
        return rows
    else:  # CSV fallback
        import csv
        with open(path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return [row["Actions"].strip() for row in reader if row.get("Actions")]


# ---------------------------------------------------------------------------
# 2. Cleaning helpers (handle OCR typos + merged lines)
# ---------------------------------------------------------------------------
ABORT_VARIANTS = ["中北登庸", "中止登廉", "中止登康"]

def normalize(text: str) -> str:
    for v in ABORT_VARIANTS:
        text = text.replace(v, "中止登庸")
    return text

def detect_ocr_variants(text: str) -> list[str]:
    """Return the list of OCR variants found in a raw row (before normalizing)."""
    return [v for v in ABORT_VARIANTS if v in text]


def load_roster(path: str) -> set[str]:
    """
    Load known officer names from a static CSV file with an 'id,name' header.
    Missing, empty, or malformed files are tolerated (returns whatever can be
    read, or an empty set) since bootstrap_names covers most names anyway.
    """
    if not os.path.exists(path):
        return set()
    names: set[str] = set()
    try:
        # utf-8-sig strips the BOM that spreadsheets often prepend
        with open(path, encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            name_col = "name" if (reader.fieldnames and "name" in reader.fieldnames) else None
            for row in reader:
                val = (row.get(name_col) if name_col else None) or ""
                val = val.strip()
                if val:
                    names.add(val)
    except OSError as e:
        print(f"  [warn] could not read roster {path}: {e}")
        return set()
    names.discard("")
    return names


def bootstrap_names(actions: list[str]) -> set[str]:
    """
    Harvest reliable names from unambiguous rows (exactly one 中止登庸,
    or a clean field/capture line). These never need roster guessing.
    """
    names: set[str] = set()
    for raw in actions:
        line = normalize(raw.strip())
        if line.count("中止登庸") == 1:
            rec, tgt = line.split("中止登庸", 1)
            names.add(rec.strip())
            names.add(tgt.strip())
        else:
            m = RE_FIELD.match(line)
            if m and "俘虜" not in line:
                names.add(m["recruiter"].strip())
                names.add(m["target"].strip())
            m = RE_CAPTURE.match(line)
            if m:
                names.add(m["target"].strip())
    names.discard("")
    return names


def _match_name(text: str, pos: int, names_by_len: list[str]) -> str | None:
    """Return the longest roster name that starts at text[pos:]."""
    for nm in names_by_len:
        if text.startswith(nm, pos):
            return nm
    return None


def split_merged(line: str, names: set[str]) -> list[str]:
    """
    Split rows that glue multiple *complete* abort events together.
    A line with fewer than two 中止登庸 markers is a single event and is
    never split. Genuine merges are cut greedily using the roster so that
    full names (關羽, 劉備, 辛毘 …) are never broken apart.

    Returns (parts, ok): `ok` is False if the line could not be parsed
    cleanly against the roster, so the caller can flag it for review.
    """
    line = normalize(line)
    if line.count("中止登庸") < 2:
        return [line] if line.strip() else []

    names_by_len = sorted(names, key=len, reverse=True)
    events, pos, n = [], 0, len(line)
    while pos < n:
        rec = _match_name(line, pos, names_by_len)
        if rec is None:
            return [line]                       # unknown name -> leave whole
        pos += len(rec)
        if not line.startswith("中止登庸", pos):
            return [line]
        pos += len("中止登庸")
        tgt = _match_name(line, pos, names_by_len)
        if tgt is None:
            return [line]
        pos += len(tgt)
        events.append(f"{rec}中止登庸{tgt}")
    return events



# ---------------------------------------------------------------------------
# 3. Regex patterns for each mechanic
# ---------------------------------------------------------------------------
RE_ABORT   = re.compile(r'^(?P<recruiter>.+?)中止登庸(?P<target>.+?)$')
RE_FIELD   = re.compile(r'^(?P<recruiter>.+?)登庸(?P<target>.+?)(?P<result>成功|失敗)$')
RE_CAPTURE = re.compile(r'^成功登庸了俘虜(?P<target>.+?)$')

def _fmt_counter(c) -> str:
    """Render a Counter of recruiters as '名前 x2、名前' (count shown if > 1)."""
    return "、".join(f"{name} x{n}" if n > 1 else name
                     for name, n in c.most_common())


def analyze(actions: list[str], names: set[str]):
    contested = Counter()
    contested_who = defaultdict(set)
    outcomes = defaultdict(lambda: {"field_success": Counter(),
                                    "field_fail": Counter()})
    issues = []

    for idx, raw in enumerate(actions, start=1):
        variants = detect_ocr_variants(raw)
        parts = split_merged(raw, names)

        marker_count = normalize(raw).count("中止登庸")
        unsplit_merge = marker_count >= 2 and len(parts) == 1
        real_merge    = len(parts) > 1

        if variants or real_merge or unsplit_merge:
            kinds = []
            if variants:
                kinds.append("OCR variant")
            if real_merge:
                kinds.append(f"merged ({len(parts)} events)")
            if unsplit_merge:
                kinds.append("UNPARSED merge - unknown name?")
            issues.append({
                "row": idx,
                "issue": " + ".join(kinds),
                "variants_found": "、".join(variants),
                "original": raw,
                "split_into": " || ".join(parts),
            })

        for line in parts:
            line = line.strip()
            # Capture lines (成功登庸了俘虜...) carry no recruiter -> skip entirely.
            if RE_CAPTURE.match(line):
                continue
            m = RE_ABORT.match(line)
            if m:
                tgt = m["target"].strip()
                contested[tgt] += 1
                contested_who[tgt].add(m["recruiter"].strip())
                continue
            m = RE_FIELD.match(line)
            if m and "俘虜" not in line:
                tgt, rec = m["target"].strip(), m["recruiter"].strip()
                if m["result"] == "成功":
                    outcomes[tgt]["field_success"][rec] += 1
                else:
                    outcomes[tgt]["field_fail"][rec] += 1

    return contested, contested_who, outcomes, issues

# ---------------------------------------------------------------------------
# 4. Console reporting (unchanged behaviour)
# ---------------------------------------------------------------------------

def report(contested, contested_who, outcomes, min_contested=2):
    print("=" * 60)
    print("SHORTLIST 1 - Contested recruitment targets (中止登庸)")
    print(f"(targets aborted by >= {min_contested} recruiters)")
    print("=" * 60)
    for tgt, cnt in contested.most_common():
        if cnt < min_contested:
            continue
        secured = bool(outcomes[tgt]["field_success"])
        landed = "later secured" if secured else "never secured"
        recruiters = "、".join(sorted(contested_who[tgt]))
        print(f"  {tgt:<6} x{cnt:<2}  [{landed}]  by: {recruiters}")

    print()
    print("=" * 60)
    print("SHORTLIST 2 - Full acquisition outcome per target")
    print("(field recruitment only)")
    print("=" * 60)
    for tgt in sorted(outcomes):
        o = outcomes[tgt]
        succ = sum(o["field_success"].values())
        fail = sum(o["field_fail"].values())
        if succ == 0 and fail == 0:
            continue
        bits = []
        if succ:
            bits.append(f"field_OK x{succ} ({'、'.join(o['field_success'])})")
        if fail:
            bits.append(f"field_FAIL x{fail} ({'、'.join(o['field_fail'])})")
        status = "JOINED" if succ else "FAILED-ONLY"
        print(f"  {tgt:<6} [{status:<11}]  " + "  |  ".join(bits))

# ---------------------------------------------------------------------------
# 5. Export everything to a new multi-sheet workbook
# ---------------------------------------------------------------------------
def _autosize(ws, max_width=80):
    """Roughly size each column to its longest cell (CJK counts as ~2)."""
    for col_cells in ws.columns:
        width = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is None:
                continue
            text = str(cell.value)
            # CJK chars are wider; weight them
            length = sum(2 if ord(ch) > 0x2E7F else 1 for ch in text)
            width = max(width, length)
        ws.column_dimensions[col_letter].width = min(width + 2, max_width)


def export_xlsx(out_path, actions, contested, contested_who,
                outcomes, issues, min_contested=2):
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    # --- Sheet 1: Original -------------------------------------------------
    ws1 = wb.active
    if ws1 is None:
        ws1 = wb.create_sheet("Original")
    else:
        ws1.title = "Original"
    ws1.append(["Row", "Actions"])
    for idx, raw in enumerate(actions, start=1):
        ws1.append([idx, raw])
    for c in ws1[1]:
        c.font = header_font

    # --- Sheet 2: Issues ---------------------------------------------------
    ws2 = wb.create_sheet("Issues")
    ws2.append(["Row", "Issue", "Variants found", "Original", "Split into"])
    for it in issues:
        ws2.append([it["row"], it["issue"], it["variants_found"],
                    it["original"], it["split_into"]])
    for c in ws2[1]:
        c.font = header_font
    if not issues:
        ws2.append(["(none)", "No OCR or merge issues detected", "", "", ""])

# --- Sheet 3: Shortlist 1 (Contested) ----------------------------------
    ws3 = wb.create_sheet("Approaching")
    ws3.append(["Target", "Abort count", "Secured?", "Secured by", "Aborted by"])
    for tgt, cnt in contested.most_common():
        if cnt < min_contested:
            continue
        o = outcomes[tgt]
        secured = bool(o["field_success"])
        ws3.append([
            tgt, cnt,
            "later" if secured else "never",
            _fmt_counter(o["field_success"]),
            "、".join(sorted(contested_who[tgt])),
        ])
    for c in ws3[1]:
        c.font = header_font

    # --- Sheet 4: Shortlist 2 (Outcomes) -----------------------------------
    ws4 = wb.create_sheet("Efforts")
    ws4.append(["Target", "Status", "Field success", "Field fail",
                "Successful recruiters", "Failed recruiters"])
    for tgt in sorted(outcomes):
        o = outcomes[tgt]
        succ = sum(o["field_success"].values())
        fail = sum(o["field_fail"].values())
        if succ == 0 and fail == 0:
            continue
        status = "JOINED" if succ else "FAILED-ONLY"
        ws4.append([
            tgt, status, succ, fail,
            _fmt_counter(o["field_success"]),
            _fmt_counter(o["field_fail"]),
        ])
    for c in ws4[1]:
        c.font = header_font




    # Cosmetic: wrap the long text columns and autosize
    for col in ("D", "E"):
        for cell in ws2[col]:
            cell.alignment = wrap
    for ws in (ws1, ws2, ws3, ws4):
        _autosize(ws)

    wb.save(out_path)
    print(f"\nExported workbook -> {out_path}")
    print(f"  Original rows : {len(actions)}")
    print(f"  Issues logged : {len(issues)}")


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    HERE = os.path.dirname(os.path.abspath(__file__))
    PATH   = os.path.join(HERE, "san14_recruitment.xlsx")
    ROSTER = os.path.join(HERE, "sangokushi14_officers.csv")
    OUT    = os.path.join(HERE, "san14_report.xlsx")

    actions = load_actions(PATH)

    # Combine the static officer roster (CSV) with names auto-harvested
    # from unambiguous rows. The union is what drives greedy splitting.
    names = load_roster(ROSTER) | bootstrap_names(actions)

    contested, contested_who, outcomes, issues = analyze(actions, names)
    report(contested, contested_who, outcomes, min_contested=2)
    export_xlsx(OUT, actions, contested, contested_who,
                outcomes, issues, min_contested=2)